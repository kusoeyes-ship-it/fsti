import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("fsti-test.html", "r") as f:
    content = f.read()

script = re.search(r'<script>(.*?)</script>', content, re.DOTALL).group(1)

q_block = re.search(r'const questions = \[(.*?)\];', script, re.DOTALL).group(1)
questions = []
for qm in re.finditer(r'q:\s*"(.*?)",\s*opts:\s*\[(.*?)\]', q_block, re.DOTALL):
    opts = []
    for om in re.finditer(r'text:\s*"(.*?)",\s*scores:\s*\{(.*?)\}', qm.group(2)):
        scores = {sm.group(1): int(sm.group(2)) for sm in re.finditer(r'(\w+):(\d+)', om.group(2))}
        opts.append({"text": om.group(1), "scores": scores})
    questions.append({"q": qm.group(1), "opts": opts})

r_block = re.search(r'const results = \[(.*?)\];', script, re.DOTALL).group(1)
results = []
for rm in re.finditer(r'code:\s*"(.*?)",\s*name:\s*"(.*?)",\s*emoji:\s*"(.*?)",\s*tagline:\s*"(.*?)",\s*dims:\s*\{(.*?)\},\s*desc:\s*"(.*?)"', r_block, re.DOTALL):
    dims = {dm.group(1): int(dm.group(2)) for dm in re.finditer(r'(\w+):(\d+)', rm.group(5))}
    results.append({"code":rm.group(1),"name":rm.group(2),"emoji":rm.group(3),"tagline":rm.group(4),"dims":dims,"desc":re.sub(r'<[^>]+>','',rm.group(6))})

# Modified content markers
mod_keywords = [
    "路人照被P得完全失真", "呼吁大家尊重真实形象", "帮忙澄清一下", "默默收藏关注",
    "手工相框——心意比金额重要", "买一件意思一下", "留言表达不满，语气坚定但保持礼貌",
    "萝卜白菜各有所爱", "审美这件事，真的没法强求", "缓一缓，晚上回家好好消化",
    "继续默默支持，陪伴不变", "又多一个人一起聊天", "提醒身边粉丝朋友注意保护爱豆隐私",
    "友好留言", "相信他/她会越来越好", "减少了关注频率", "多听几遍",
    "好歌自己会说话", "官方物料、演出照", "好看的照片和有趣的视频",
    "一直默默牵挂着", "编外小助手", "勤劳小蜜蜂", "应援达人", "淡定老粉",
    "专一守护者", "应援组织者", "慷慨应援粉", "今天你签到了吗",
    "为爱豆花的每一分钱", "新物料出了！一起来讨论", "签到我来、应援我来",
    "不折腾", "不刷屏、不争论、不跟风", "天天签到", "逢帖必回",
    "关注动态，晚上整理物料", "理性消费，快乐追星", "追星在于真心而非金额",
    "耐心地用事实进行澄清", "温和但坚定地解释", "脑内剧场永远在上映",
    "组织大家讨论", "张罗应援", "粉丝社区", "参加过最早的签售",
    "应援物和满格电的手机", "量力而行",
]

def is_modified(text):
    return any(kw in text for kw in mod_keywords)

wb = Workbook()
hfont = Font(bold=True, color="FFFFFF", size=11)
rfont = Font(color="FF0000", bold=True, size=10)
rfill = PatternFill("solid", fgColor="FFF0F0")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(left=Side(style="thin",color="D4D4D4"),right=Side(style="thin",color="D4D4D4"),top=Side(style="thin",color="D4D4D4"),bottom=Side(style="thin",color="D4D4D4"))

# Sheet 1
ws1 = wb.active
ws1.title = "题目清单"
for c, h in enumerate(["题号","题目","选项A","A计分","选项B","B计分","选项C","C计分","选项D","D计分"], 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = PatternFill("solid", fgColor="7C3AED"); cell.alignment = Alignment(horizontal="center",vertical="center"); cell.border = bdr

for i, q in enumerate(questions):
    row = i + 2
    ws1.cell(row=row, column=1, value=i+1).alignment = Alignment(horizontal="center",vertical="top"); ws1.cell(row=row,column=1).border = bdr
    qc = ws1.cell(row=row, column=2, value=q["q"]); qc.alignment = wrap; qc.border = bdr
    if is_modified(q["q"]): qc.font = rfont; qc.fill = rfill
    for j, opt in enumerate(q["opts"]):
        oc = ws1.cell(row=row, column=3+j*2, value=opt["text"]); oc.alignment = wrap; oc.border = bdr
        if is_modified(opt["text"]): oc.font = rfont; oc.fill = rfill
        sc = ws1.cell(row=row, column=4+j*2, value=", ".join(f"{k}:{v}" for k,v in opt["scores"].items())); sc.alignment = wrap; sc.border = bdr

ws1.column_dimensions["A"].width = 5; ws1.column_dimensions["B"].width = 40
for cl in ["C","E","G","I"]: ws1.column_dimensions[cl].width = 45
for cl in ["D","F","H","J"]: ws1.column_dimensions[cl].width = 18

# Sheet 2
ws2 = wb.create_sheet("人格结果")
dn = {"krypton":"应援力","data":"数据力","face":"颜控度","mom":"妈粉值","gf":"恋爱脑","toxic":"专一度","chill":"淡定度","climb":"爬墙率","shill":"安利欲","antihate":"反黑力","ctrl":"组织力","front":"前线力","drama":"戏精值","emo":"emo值","sane":"理智值","loyal":"忠诚度"}
ad = list(dn.keys())
for c, h in enumerate(["序号","代号","Emoji","人格名称","标志台词"]+[dn[d] for d in ad]+["人格描述"], 1):
    cell = ws2.cell(row=1, column=c, value=h); cell.font = hfont; cell.border = bdr
    cell.fill = PatternFill("solid", fgColor="EC4899") if c<=5 else (PatternFill("solid", fgColor="7C3AED") if c<=5+len(ad) else PatternFill("solid", fgColor="3B82F6"))
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

for i, r in enumerate(results):
    row = i + 2
    ws2.cell(row=row,column=1,value=i+1).alignment=Alignment(horizontal="center",vertical="top"); ws2.cell(row=row,column=1).border=bdr
    ws2.cell(row=row,column=2,value=r["code"]).border=bdr
    ws2.cell(row=row,column=3,value=r["emoji"]).alignment=Alignment(horizontal="center"); ws2.cell(row=row,column=3).border=bdr
    nc = ws2.cell(row=row,column=4,value=r["name"]); nc.border=bdr
    if is_modified(r["name"]): nc.font=Font(bold=True,color="FF0000"); nc.fill=rfill
    else: nc.font=Font(bold=True)
    tc = ws2.cell(row=row,column=5,value=r["tagline"]); tc.alignment=wrap; tc.border=bdr
    if is_modified(r["tagline"]): tc.font=rfont; tc.fill=rfill
    for di, dk in enumerate(ad):
        val = r["dims"].get(dk,"")
        cell = ws2.cell(row=row,column=6+di,value=val if val else ""); cell.alignment=Alignment(horizontal="center",vertical="top"); cell.border=bdr
        if val and val>=80: cell.font=Font(bold=True,color="7C3AED")
    dc = ws2.cell(row=row,column=6+len(ad),value=r["desc"]); dc.alignment=wrap; dc.border=bdr
    if is_modified(r["desc"]): dc.font=rfont; dc.fill=rfill

ws2.column_dimensions["A"].width=5; ws2.column_dimensions["B"].width=10; ws2.column_dimensions["C"].width=5; ws2.column_dimensions["D"].width=16; ws2.column_dimensions["E"].width=35
for di in range(len(ad)): ws2.column_dimensions[get_column_letter(6+di)].width=8
ws2.column_dimensions[get_column_letter(6+len(ad))].width=60

# Sheet 3: 整改对照
ws3 = wb.create_sheet("安全整改对照")
for c, h in enumerate(["序号","安全意见","修改位置","修改前（旧）","修改后（新）"], 1):
    cell = ws3.cell(row=1,column=c,value=h); cell.font=hfont; cell.fill=PatternFill("solid",fgColor="EF4444"); cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=bdr

changes = [
    ("1","伪造商品好评","Q4选项D","去各大平台刷好评冲销量，精神氪金也是氪金","默默收藏关注，在心里为爱豆加油就好"),
    ("2","大额现金赠与","Q7选项B","直接包了个五位数的现金红包","花两周亲手做了手工相框——心意比金额重要"),
    ("3","代拍乱象","Q2题目+选项","代拍图/代拍行业乱象/饭拍代拍路透","路人照/尊重真实形象/官方物料演出照"),
    ("4","不良追星导向","Q2选项D","姐妹们上号！一人十条给我冲","大家来看看原图，帮忙澄清一下吧"),
    ("5","超前消费","Q4选项B","算了不算了，先买再说","买一件意思一下吧"),
    ("6","饭圈洗脑/盲从","Q18选项A","无条件支持！数据照做！安利照发！","虽然有点失落，但相信他/她会越来越好"),
    ("7","境外竞品网站","CP脑结果描述","浏览器永远开着AO3和LOFTER","脑内剧场永远在上映"),
    ("8","刷屏/打投/控评","多处选项+描述","打投/控评/刷播放量/做数据","签到/关注/讨论/应援/多听几遍"),
    ("9","饭圈互撕","Q10选项A+B","踢了踢了！间谍/截图存证以备后用","萝卜白菜各有所爱/审美没法强求"),
    ("10","追星不良价值观","Q13选项A","请三天假来处理情绪","缓一缓，晚上回家好好消化"),
    ("11","号召抵制=网暴","Q16选项A","号召全网抵制","提醒身边粉丝保护爱豆隐私"),
    ("12","分类名不良导向","结果名称×6","氪金战神/控评总指挥/数据女工/毒唯/佛系/富婆富哥","应援达人/应援组织者/勤劳小蜜蜂/专一守护者/淡定老粉/慷慨应援粉"),
    ("13","氪金不良导向","应援达人描述","真金白银表达爱意/信用卡账单是勋章","理性消费，快乐追星/一切都值得"),
    ("14","宗教敏感","维度名+结果名","佛系度/佛系老粉","淡定度/淡定老粉"),
    ("15","炫富导向","慷慨应援粉描述","有钱人的追星快乐你想象不到","追星在于真心而非金额"),
]

for i, (num,opinion,loc,old,new) in enumerate(changes):
    row = i + 2
    ws3.cell(row=row,column=1,value=num).alignment=Alignment(horizontal="center",vertical="top"); ws3.cell(row=row,column=1).border=bdr
    ws3.cell(row=row,column=2,value=opinion).alignment=wrap; ws3.cell(row=row,column=2).border=bdr
    ws3.cell(row=row,column=3,value=loc).alignment=wrap; ws3.cell(row=row,column=3).border=bdr
    oc = ws3.cell(row=row,column=4,value=old); oc.alignment=wrap; oc.font=Font(color="999999",strikethrough=True); oc.border=bdr
    nc = ws3.cell(row=row,column=5,value=new); nc.alignment=wrap; nc.font=Font(color="FF0000",bold=True); nc.fill=rfill; nc.border=bdr

ws3.column_dimensions["A"].width=5; ws3.column_dimensions["B"].width=18; ws3.column_dimensions["C"].width=18; ws3.column_dimensions["D"].width=45; ws3.column_dimensions["E"].width=45

ws1.freeze_panes="C2"; ws2.freeze_panes="F2"; ws3.freeze_panes="A2"

out = "FSTI测试内容导出（安全整改版）.xlsx"
wb.save(out)
print(f"Saved! Q={len(questions)} R={len(results)}")
